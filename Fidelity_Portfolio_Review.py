import requests
import re
from openpyxl import load_workbook
import time

# change the file name to the file name of your workbook
# at this time, the source workbook must be an .xlsx file
source_workbook = load_workbook(filename = 'your_fidelity_portfolio.xlsx')
# change the name below to your sheet name
source_portfolio = source_workbook['portfolio_sheet_name']

# load in workbook for where you want to print to
target_workbook = load_workbook(filename = 'target_workbook.xlsx')
target_portfolio = target_workbook['Portfolio']

def get_info(fetch):
    # classification find
    if fetch == 'classification':
        classification_location = data.find('classification')
        classification = data[(classification_location + 17) : (classification_location + 30)]
        print(classification)
        return classification
    
    # ETF PB Ratio
    if fetch == 'etf_pb_ratio':
        pb_ratio_location = data.find('priceToBook')
        pb_ratio_rough = data[(pb_ratio_location + 12) : (pb_ratio_location + 22)]
        pb_ratio = re.findall(r"[-+]?\d*\.\d+|\d+", pb_ratio_rough)
        return pb_ratio
    
    # ETF PS Ratio
    if fetch == 'etf_ps_ratio':
        ps_ratio_location = data.find('priceToSales')
        ps_ratio_rough = data[(ps_ratio_location + 13) : (ps_ratio_location + 23)]
        ps_ratio = re.findall(r"[-+]?\d*\.\d+|\d+", ps_ratio_rough)
        return ps_ratio
    
    # Instit Ownership (stock and etf)
    if fetch == 'instit_ownership':
        instit_ownership_location = data.find('institOwnership')
        instit_ownership_rough = data[(instit_ownership_location + 16) : (instit_ownership_location + 26)]
        instit_ownership = re.findall(r"[-+]?\d*\.\d+|\d+", instit_ownership_rough)
        return instit_ownership
    
    # market return month-end one year (ETF)
    if fetch == 'etf_one_year_return':
        one_year_return_location = data.find('marketReturnMonthEnd1Yr')
        rough_one_year_return = data[(one_year_return_location + 24) : (one_year_return_location + 34)]
        one_year_return = re.findall(r"[-+]?\d*\.\d+|\d+", rough_one_year_return)
        return one_year_return
    
    # equity summary detail
    if fetch == 'equity_summary_detail':
        equity_summary_location = data.find('equitySummDetail')
        rough_equity_summary = data[(equity_summary_location + 25) : (equity_summary_location + 33)]
        equity_summary_score = re.findall(r"[-+]?\d*\.\d+|\d+", rough_equity_summary)
        return equity_summary_score
    
    # Recognia Short Term Direction
    if fetch == 'recognia_short_term':
        recognia_short_term_location = data.find('shortTermDirection')
        rough_short_term_data = data[recognia_short_term_location : (recognia_short_term_location + 29)]
        if 'Bearish' in rough_short_term_data:
            recognia_short_term = 'Bearish'
        elif 'Bullish' in rough_short_term_data:
            recognia_short_term = 'Bullish'
        elif 'Neutral' in rough_short_term_data:
            recognia_short_term = 'Neutral'
        else:
            recognia_short_term = 'Unavailable'
        return recognia_short_term
    
    # Recognia Intermediate Term Direction
    if fetch == 'recognia_intermediate_term':
        recognia_intermediate_term_location = data.find('intermediateTermDirection')
        rough_intermediate_term_data = data[recognia_intermediate_term_location : (recognia_intermediate_term_location + 36)]
        if 'Bearish' in rough_intermediate_term_data:
            recognia_intermediate_term = 'Bearish'
        elif 'Bullish' in rough_intermediate_term_data:
            recognia_intermediate_term = 'Bullish'
        elif 'Neutral' in rough_intermediate_term_data:
            recognia_intermediate_term = 'Neutral'
        else:
            recognia_intermediate_term = 'Unavailable'
        return recognia_intermediate_term
    
    # Recognia Long Term Direction
    if fetch == 'recognia_long_term':
        recognia_long_term_location = data.find('longTermDirection')
        rough_long_term_data = data[recognia_long_term_location : (recognia_long_term_location + 27)]
        if 'Bearish' in rough_long_term_data:
            recognia_long_term = 'Bearish'
        elif 'Bullish' in rough_long_term_data:
            recognia_long_term = 'Bullish'
        elif 'Neutral' in rough_long_term_data:
            recognia_long_term = 'Neutral' 
        else:
            recognia_long_term = 'Unavailable'
        return recognia_long_term

    # PE Ratio (stock)
    if fetch == 'pe_ratio':
        pe_ratio_location = data.find('peCurrYrEst')
        rough_pe_ratio = data[(pe_ratio_location + 12) : (pe_ratio_location + 23)]
        pe_ratio = re.findall(r"[-+]?\d*\.\d+|\d+", rough_pe_ratio)
        return pe_ratio
    
    # Price to Book TTM
    if fetch == 'pb_ratio':
        pb_ratio_location = data.find('priceToBookTTM')
        rough_pb_ratio = data[(pb_ratio_location + 15) : (pb_ratio_location + 27)]
        pb_ratio = re.findall(r"[-+]?\d*\.\d+|\d+", rough_pb_ratio)
        return pb_ratio
    
    # Price to Sales TTM
    if fetch == 'ps_ratio':
        ps_ratio_location = data.find('pricetoSalesTTM')
        rough_ps_ratio = data[(ps_ratio_location + 15) : (ps_ratio_location +27)]
        ps_ratio = re.findall(r"[-+]?\d*\.\d+|\d+", rough_ps_ratio)
        return ps_ratio
    
    # Debt to Equity
    if fetch == 'debt_to_equity_ratio':
        debt_to_equity_location = data.find('longTermDebtToEquityTTM')
        rough_debt_to_equity = data[(debt_to_equity_location + 23) : (debt_to_equity_location + 37)]
        debt_to_equity_ratio = re.findall(r"[-+]?\d*\.\d+|\d+", rough_debt_to_equity)
        return debt_to_equity_ratio

    # total one year return
    if fetch == 'one_year_return':
        one_year_return_location = data.find('totalReturn1Yr')
        rough_one_year_return = data[(one_year_return_location + 15) : (one_year_return_location + 26)]
        one_year_return = re.findall(r"[-+]?\d*\.\d+|\d+", rough_one_year_return)
        return one_year_return

i = 2

# change the max_row amount to the number of rows in your portfolio excel document
for row in source_portfolio.iter_rows(min_row=2, max_col=14, max_row=154, values_only=True):
    account, symbol, description, quantity, last_price, last_price_change, current_value, todays_gain_loss_dollar, todays_gain_loss_percent, total_gain_loss_dollar, total_gain_loss_percent, percent_of_account, cost_basis, cost_basis_per_share = row

    # call cells to write to in target workbook
    stock_symbol = target_portfolio.cell(row=i, column=1)
    stock_description = target_portfolio.cell(row=i, column=2)
    stock_quantity = target_portfolio.cell(row=i, column=3)
    stock_last_price = target_portfolio.cell(row=i, column=4)
    stock_last_price_change = target_portfolio.cell(row=i, column=5)
    stock_current_value = target_portfolio.cell(row=i, column=6)
    stock_todays_gain_loss_dollar = target_portfolio.cell(row=i, column=7)
    stock_todays_gain_loss_percent = target_portfolio.cell(row=i, column=8)
    stock_total_gain_loss_dollar = target_portfolio.cell(row=i, column=9)
    stock_total_gain_loss_percent = target_portfolio.cell(row=i, column=10)
    stock_percent_of_account = target_portfolio.cell(row=i, column=11)
    stock_cost_basis = target_portfolio.cell(row=i, column=12)
    stock_cost_basis_per_share = target_portfolio.cell(row=i, column=13)
    stock_equity_summary_score = target_portfolio.cell(row=i, column=14)
    stock_recognia_short_term = target_portfolio.cell(row=i, column=15)
    stock_recognia_intermediate_term = target_portfolio.cell(row=i, column=16)
    stock_recognia_long_term = target_portfolio.cell(row=i, column=17)
    # skip columns 18 and 19 as I can't automate those
    stock_pe_ratio = target_portfolio.cell(row=i, column=20)
    stock_pb_ratio = target_portfolio.cell(row=i, column=21)
    stock_ps_ratio = target_portfolio.cell(row=i, column=22)
    stock_debt_to_equity_ratio = target_portfolio.cell(row=i, column=23)
    stock_one_year_return = target_portfolio.cell(row=i, column=24)
    stock_instit_ownership = target_portfolio.cell(row=i, column=25)
    stock_acct = target_portfolio.cell(row=i, column=30)

    # copy static data
    stock_symbol.value = symbol
    stock_description.value = description
    stock_quantity.value = quantity
    stock_last_price.value = last_price
    stock_last_price_change.value = last_price_change
    stock_current_value.value = current_value
    stock_todays_gain_loss_dollar.value = todays_gain_loss_dollar
    stock_todays_gain_loss_percent.value = todays_gain_loss_percent
    stock_total_gain_loss_dollar.value = total_gain_loss_dollar
    stock_total_gain_loss_percent.value = total_gain_loss_percent
    stock_percent_of_account.value = percent_of_account
    stock_cost_basis.value = cost_basis
    stock_cost_basis_per_share.value = cost_basis_per_share
    stock_acct.value = account

    prev_stock_symbol = target_portfolio.cell(row=(i - 1), column=1)
    print('Prev symbol is: ' + str(prev_stock_symbol))
    print('Current symbol is ' + str(symbol))
    if str(prev_stock_symbol.value) == str(symbol):
        print("copying data for duplicate symbol" + symbol)
        prev_stock_equity_summary_score = target_portfolio.cell(row=(i-1), column=14)
        prev_stock_recognia_short_term = target_portfolio.cell(row=(i-1), column=15)
        prev_stock_recognia_intermediate_term = target_portfolio.cell(row=(i-1), column=16)
        prev_stock_recognia_long_term = target_portfolio.cell(row=(i-1), column=17)
        # skip columns 18 and 19 as I can't automate those
        prev_stock_pe_ratio = target_portfolio.cell(row=(i-1), column=20)
        prev_stock_pb_ratio = target_portfolio.cell(row=(i-1), column=21)
        prev_stock_ps_ratio = target_portfolio.cell(row=(i-1), column=22)
        prev_stock_debt_to_equity_ratio = target_portfolio.cell(row=(i-1), column=23)
        prev_stock_one_year_return = target_portfolio.cell(row=(i-1), column=24)
        prev_stock_instit_ownership = target_portfolio.cell(row=(i-1), column=25)

        # copy data
        stock_equity_summary_score.value = prev_stock_equity_summary_score.value
        stock_recognia_short_term.value = prev_stock_recognia_short_term.value
        stock_recognia_intermediate_term.value = prev_stock_recognia_intermediate_term.value
        stock_recognia_long_term.value = prev_stock_recognia_long_term.value
        # skip columns 18 and 19 as I can't automate those
        stock_pe_ratio.value = prev_stock_pe_ratio.value
        stock_pb_ratio.value = prev_stock_pb_ratio.value
        stock_ps_ratio.value = prev_stock_ps_ratio.value
        stock_debt_to_equity_ratio.value = prev_stock_debt_to_equity_ratio.value
        stock_one_year_return.value = prev_stock_one_year_return.value
        stock_instit_ownership.value = prev_stock_instit_ownership.value

        print("copied data for " + symbol)
    else:
        
        print('performing API call for ' + symbol)

        #commenting out API calls until I'm sure my syntax is OK
        url = "https://fidelity-investments.p.rapidapi.com/quotes/get-mashup"

        querystring = {"symbol":symbol}

        headers = {
            'x-rapidapi-host': "fidelity-investments.p.rapidapi.com",
            'x-rapidapi-key': "your RapidAPI Key goes here"
        } 

        response = requests.request("GET", url, headers=headers, params=querystring)

        data = response.text
        print('api call performed')

        if 'Service Unavailable' in data:
            print('API Service Unavailable. Check down status on Rapid API or your internet connection.')
            exit()

        # Parse API call data and copy it into excel

        classification = get_info(fetch = 'classification')
        if 'ETF' in classification:
            print(symbol + 'is ETF')

            # etf PB ratio
            pb_ratio = get_info('etf_pb_ratio')
            if pb_ratio == []:
                stock_pb_ratio.value = 'NA'
            else:
                stock_pb_ratio.value = float(pb_ratio[0])
            
            # etf ps ratio
            ps_ratio = get_info('etf_ps_ratio')
            if ps_ratio == []:
                stock_ps_ratio.value = 'NA'
            else:
                stock_ps_ratio.value = float(ps_ratio[0])
            
            # etf instit ownership
            instit_ownership = get_info('instit_ownership')
            if instit_ownership == []:
                stock_instit_ownership.value = 'NA'
            else:
                stock_instit_ownership.value = float(instit_ownership[0])
            
            # etf one year return
            one_year_return = get_info('etf_one_year_return')
            if one_year_return == []:
                stock_one_year_return.value = 'NA'
            else:
                stock_one_year_return.value = float(one_year_return[0])
            
            print(symbol + 'complete')
            
        else:
            print(symbol + 'is something else')

            equity_summary_score = get_info(fetch = 'equity_summary_detail')
            if equity_summary_score == []:
                stock_equity_summary_score.value = 'NA'
            else:
                stock_equity_summary_score.value = float(equity_summary_score[0])
        
            recognia_short_term = get_info('recognia_short_term')
            stock_recognia_short_term.value = recognia_short_term
            
            recognia_intermediate_term = get_info('recognia_intermediate_term')
            stock_recognia_intermediate_term.value = recognia_intermediate_term
            
            recognia_long_term = get_info('recognia_long_term')
            stock_recognia_long_term.value = recognia_long_term

            pe_ratio = get_info('pe_ratio')
            if pe_ratio == []:
                stock_pe_ratio.value = 'NA'
            else:
                stock_pe_ratio.value = float(pe_ratio[0])
            
            pb_ratio = get_info('pb_ratio')
            if pb_ratio == []:
                stock_pb_ratio.value = 'NA'
            else:
                stock_pb_ratio.value = float(pb_ratio[0])
            
            ps_ratio = get_info('ps_ratio')
            if ps_ratio == []:
                stock_ps_ratio.value = 'NA'
            else:
                stock_ps_ratio.value = float(ps_ratio[0])
            
            debt_to_equity_ratio = get_info('debt_to_equity_ratio')
            if debt_to_equity_ratio == []:
                stock_debt_to_equity_ratio.value = 'NA'
            else:
                stock_debt_to_equity_ratio.value = float(debt_to_equity_ratio[0])
            
            instit_ownership = get_info('instit_ownership')
            if instit_ownership == []:
                stock_instit_ownership.value = 'NA'
            else:
                stock_instit_ownership.value = float(instit_ownership[0])

            one_year_return = get_info('one_year_return')
            if one_year_return == []:
                stock_one_year_return.value = 'NA'
            else:
                stock_one_year_return.value = float(one_year_return[0])
            
            print('printed data for ' + symbol)
        
    i += 1

    # save workbook every ten rows
    if (i % 10) == 0:
        target_workbook.save('target_workbook.xlsx')
        print('file saved')
    
    # delay to help w/ API calls
    time.sleep(1)

print('saving workbook...')
target_workbook.save('target_workbook.xlsx')
print('Workbook Saved')
print('Done')
